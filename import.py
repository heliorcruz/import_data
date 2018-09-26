import pymongo
from openpyxl import load_workbook
import config as cf


def import_data():
    '''
       Import data from .xlsx and store to mongo
    '''    
    wb = load_workbook(filename = cf.FILENAME)
    # To display all of the available worksheet names
    sheets = wb.sheetnames
    print sheets
    # To work with the first sheet (by name)
    ws = wb[sheets[0]]
    print "\n"
    
    CLIENT_LIST = []
    count = 1
    for row in range(cf.HEADER,cf.ROWS):
        client_data = {}
        categ_data  = []
        cel_data    = ""
        
        for col in range(0,cf.COLLUMS):
            sheet_col = cf.ALFA[col]+str(row + 1)
            key_data = cf.FIELDS[col]
            col_data = ws[sheet_col].value
            if key_data in CATEG:
                cl_data = col_data.encode('utf-8') if type(col_data) in(str,unicode) else str(col_data)
                if cl_data != 'None' and cl_data != '':
                    categ_data.append(cl_data)       
                    
            else:
                cl_data = col_data.encode('utf-8') if type(col_data) in(str,unicode) else str(col_data)
                client_data[key_data] = cl_data if cl_data != 'None' else ''

        client_data[cf.CATEG[0]]  = categ_data
        client_data[cf.PHONES[1]] = cel_data        
        CLIENT_LIST.append(client_data)  
        print str(count) +"  "+ str(client_data)        
        count += 1        
        
           
    dbCon = pymongo.MongoClient(cf.CONN_STRING)
    db = dbCon[cf.DATABASE]    
    client = db[cf.COLLECTION]
    
    count = 1
    for data in CLIENT_LIST:
        if cf.UPSERT:
            ret = client.update_one( data ,{ '$set':  data   }, upsert=True)
            print str(count)+ " documents verified"
        else:
            ret  = client.insert_one(data)        
            print str(count)+ " key: "+ str(ret.inserted_id)
            
        count += 1
        
    print "Total {0} ".format("INSERTED" if not cf.UPSERT else "UPDATED")+str(count)
    
   
def process_duplicates():
    '''
       Process duplicate records in tables
    '''
    
    print "\n"        
           
    dbCon = pymongo.MongoClient(cf.CONN_STRING)
    db = dbCon[cf.DATABASE]    
    client = db[cf.COLLECTION]
    # FIRST  {"$match": {"DATA_NASC":{"$ne":""}}}   SECOND  {"$match": {"DATA_NASC":""}}
    ALL_DATA = client.aggregate([
                                       {                                            
                                         "$group": { 
                                         "_id": { "NOME": "$NOME", "DATA_NASC":"$DATA_NASC" }, 
                                         "uniqueIds": { "$addToSet": "$_id" },
                                         "count": { "$sum": 1 } 
                                        }}, 
                                          { "$match": { 
                                            "count": { "$gt": 1 } 
                                        }}
                                     ])
    count = 1
    tot_count = 0
    for data in ALL_DATA:
        print "\n--------------------------------------------------------------------"
        print str(data["_id"])+"   Registros encontrados: "+str(data["count"])
        count = 1
        unique_ids = [str(unik) for unik in data["uniqueIds"]]
        for info in data["uniqueIds"]:
            ret  = client.find_one({"_id":info})
            print str(count)+ " key: "+ str(ret["_id"])+ " " + ret["NOME"]            
            duplicated_ids = [unik for unik in unique_ids]
            duplicated_ids.remove(str(ret["_id"]))
            client.update({"_id":info}, {"$set": {"IS_DUPLICATED": duplicated_ids}})
            count += 1                   
            
        tot_count += 1
        
    print "Total duplicates "+str(tot_count)
    

def run():
    print "Executando script..."
    try:
        import_data()               
        process_duplicates()        
    except Exception as wE:
        print wE
    print "Fim da importacao."   

run();


